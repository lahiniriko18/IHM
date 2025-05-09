from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime
from ..serializer.serializerEdt import EdtSerializer
from ..serializer.serializerPosseder import PossederSerializer
from ..serializer.serializerConstituer import ConstituerSerializer
from ..serializer.serializerExcel import ExcelSerializer,DataSerializer
import pandas as pd
from openpyxl import load_workbook
from ..models import Edt
class EdtView(APIView):
    def get(self, request):
        edts=Edt.objects.all()
        serializer=EdtSerializer(edts, many=True)
        return Response(serializer.data)
    
    def post(self, request):
        serializer=EdtSerializer(data=request.data)
        if serializer.is_valid():
            edt=serializer.save()
            return Response(EdtSerializer(edt).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    def put(self, request, numEdt):
        try:
            edt=Edt.objects.get(pk=numEdt)
        except Edt.DoesNotExist:
            return Response({"erreur":"Emploi du temps introuvable"}, status=status.HTTP_404_NOT_FOUND)
        serializer=EdtSerializer(edt, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    def delete(self, request, numEdt):
        try:
            edt=Edt.objects.get(pk=numEdt)
            edt.delete()
            return Response({'message':'Suppression avec succès'}, status=status.HTTP_200_OK)
        except Edt.DoesNotExist:
            return Response({'erreur':'Emploi du temps introuvable'}, status=status.HTTP_404_NOT_FOUND)

class EdtTableauView(APIView):
    def post(self, request):
        donnees=request.data
        donneeAjout = []
        if isinstance(donnees, list):
            return Response({"erreur":"Format de données invalide"})
        for donnee in donnees:
            serializer=EdtSerializer(data=donnee)
            if serializer.is_valid():
                edt=serializer.save()
                donneeAjout.append(EdtSerializer(edt).data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        return Response(donneeAjout,status=status.HTTP_201_CREATED)

class EdtExcelView(APIView):
    def post(self, request):
        serializer=ExcelSerializer(data=request.data)

        if serializer.is_valid():
            fichier=serializer.validated_data['fichier']
            typeFichier=int(serializer.validated_data["typeFichier"])
            try:
                wb = load_workbook(fichier)
                ws = wb.active
                ligneWb=list(ws.iter_rows(values_only=True))
                premierLignes=pd.read_excel(fichier, nrows=2)
                
                colonnes_requis=['Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi']
                jours=['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi']
                premierLignes.columns=['Titre']
                
                colUtile=[]
                dataUtile = []
                if len(ligneWb)<5:
                    return Response(
                        {
                            "texte":"Format invalide !",
                            "aide":"Colonne requis: 'Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'"
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
                if typeFichier==1:
                    verif=0
                    for i,col in enumerate(ligneWb[3]):
                        if i<13:
                            if i== 0 and col is None:
                                colUtile.append("Horaire")
                            elif col is None:
                                colUtile.append(f"Unnamed: {i}")
                            else:
                                colUtile.append(col.strip())
                                verif+=1
                    if verif!=6:
                        return Response(
                            {
                                "texte":"Format invalide !",
                                "aide":"Colonne requis: 'Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'"
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    for i,lignes in enumerate(ligneWb[4:]):
                        ligne = []
                        for j,v in enumerate(lignes):
                            if j<13:
                                ligne.append(v.strip() if v else "vide")
                        dataUtile.append(ligne)
                
                else:
                    horairesLigne=[]
                    for i,col in enumerate(ligneWb[3]):
                        if col and i<13:
                            horairesLigne.append(col)
                    joursLigne=[]
                    for i,lignes in enumerate(ligneWb[4:]):
                        joursLigne.append(lignes[0].strip())
                    if len(joursLigne) != 6:
                        return Response(
                            {
                                "texte":"Ligne jour manquant !",
                                "aide":"Ligne requis: 'Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'"
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    colUtile=["Horaire"]
                    for j,jour in enumerate(joursLigne):
                        colUtile.append(jour)
                        colUtile.append(f"Unnamed {j}")
                    print(colUtile)
                    for i,horaire in enumerate(ligneWb[3]):
                        verif = (i<4 and i%2!=0) or (i>5 and i%2==0)
                        if verif and i<9:
                            ligneDonne=[horaire]
                            for lignes in ligneWb[4:]:
                                ligneDonne.append(lignes[i] if lignes[i] else "vide")
                                ligneDonne.append(lignes[i+1] if lignes[i+1] else "vide")
                            dataUtile.append(ligneDonne)   
                        
                df=pd.DataFrame(dataUtile, columns=colUtile)

                if not all(col in df.columns for col in colonnes_requis):
                    return Response({"erreur":"Format invalide. Colonne requis: 'Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'"}, status=status.HTTP_400_BAD_REQUEST)
                colonneVerif=[col for col in df.columns if col in colonnes_requis]
                for i,v in enumerate(colonneVerif):
                    if v != colonnes_requis[i]:
                        return Response({"erreur":"Format invalide. Colonne requis: 'Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'"}, status=status.HTTP_400_BAD_REQUEST)
                        
                dataContenu=[]
                dataTitre=premierLignes.to_dict(orient='records')
                colonnes=df.columns
                
                for i,ligne in df.iterrows():
                    lignes = {"Horaire":ligne['Horaire']}
                    if ligne['Horaire'] != "vide":
                        j=1
                        while j<len(colonnes):
                            jour=colonnes[j]
                            valeurs=[ligne.get(jour)]
                            k=1
                            while j+k< len(colonnes) and colonnes[j+k] not in jours:
                                valeurs.append(ligne.get(colonnes[j+k]))
                                k+=1
                            j+=k
                            lignes[jour]=valeurs
                        dataContenu.append(lignes)
                data={
                    "titre":dataTitre,
                    "contenu":dataContenu,
                    "typeFichier":typeFichier
                }
                # test git hub
                serializer=DataSerializer(data=data)
                if serializer.is_valid():
                    donnee = serializer.validated_data
                    contenu=donnee["contenu"]
                    dates=donnee["titre"][0]
                    classeParcours=donnee["titre"][1]
                    if len(contenu) > 1:
                        heureCourant=contenu[0]["Horaire"]["heureFin"]
                        for i in range(1,len(contenu)):
                            if heureCourant > contenu[i]["Horaire"]["heureDebut"]:
                                return Response({"erreur":f"L'heure de fin dans la ligne {i} doit inférieure ou égale à l'heure de début de la ligne {i+1} dans le colonne de horaire !"})
                            heureCourant=contenu[i]["Horaire"]["heureFin"]
                    if len(contenu) > 0:
                        if not classeParcours["constitue"]:
                            donneeConstituer={
                                "numParcours":classeParcours["parcours"],
                                "numClasse":classeParcours["classe"]
                            }
                            serializerConstituer = ConstituerSerializer(data=donneeConstituer)
                            if serializerConstituer.is_valid():
                                serializerConstituer.save()
                        for ligne in contenu:
                            horaire=ligne["Horaire"]
                            for jour in jours:
                                valeurJour=ligne.get(jour)
                                for val in valeurJour:
                                    if isinstance(val, dict):
                                        if not val["classeGroupe"]:
                                            donneePosseder={
                                                "numClasse":classeParcours["classe"],
                                                "numGroupe":val["groupe"]
                                            }
                                            serializerPosseder=PossederSerializer(data=donneePosseder)
                                            if serializerPosseder.is_valid():
                                                serializerPosseder.save()
                                        dateObj=datetime.strptime(dates.get(jour.lower()),"%d-%m-%Y")
                                        dateSql=dateObj.strftime("%Y-%m-%d")
                                        donneEdt={
                                            "numMatiere":val["matiere"],
                                            "numParcours":classeParcours["parcours"],
                                            "numSalle":val["salle"],
                                            "numClasse":classeParcours["classe"],
                                            "date":dateSql,
                                            "heureDebut":horaire["heureDebut"],
                                            "heureFin":horaire["heureFin"]
                                        }
                                        serializerEdt= EdtSerializer(data=donneEdt)
                                        if serializerEdt.is_valid():
                                            serializerEdt.save()
                                        else:
                                            return Response(serializerEdt.errors, status=status.HTTP_400_BAD_REQUEST)
                    return Response({"message":"Fichier traité et ajout d'emploi du temps avec succès !"}, status=status.HTTP_200_OK)
                return Response(serializer.errors,status=status.HTTP_401_UNAUTHORIZED)
            except Exception as e:
                return Response(str(e), status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
