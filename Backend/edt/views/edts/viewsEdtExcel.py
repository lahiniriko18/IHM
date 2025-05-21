from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from ...serializer.serializerExcel import ExcelSerializer,DataSerializer
import pandas as pd
from openpyxl import load_workbook
from ...services.serviceEdt import ServiceEdtCrud


class EdtExcelView(APIView):
    def post(self, request):
        serializer=ExcelSerializer(data=request.data)

        if serializer.is_valid():
            fichier=serializer.validated_data['fichier']
            typeFichier=int(serializer.validated_data["typeFichier"])
            try:
                wb = load_workbook(fichier)
                ws = wb.active
                ligneWb=list(ws.iter_rows(values_only=True))
                premierLignes=pd.read_excel(fichier, nrows=2)
                
                colonnes_requis=['Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi']
                jours=['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi']
                premierLignes.columns=['Titre']
                
                colUtile=[]
                dataUtile = []
                if len(ligneWb)<5:
                    return Response(
                        {
                            "texte":"Format invalide !",
                            "aide":"Colonne requis: 'Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'"
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
                if typeFichier==1:
                    verif=0
                    for i,col in enumerate(ligneWb[3]):
                        if i== 0 and col is None:
                            colUtile.append("Horaire")
                        elif col is None:
                            colUtile.append(f"Unnamed: {i}")
                        else:
                            colUtile.append(col.strip())
                            verif+=1
                    if verif!=6:
                        return Response(
                            {
                                "texte":"Format invalide !",
                                "aide":"Colonne requis: 'Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'"
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    for i,lignes in enumerate(ligneWb[4:]):
                        ligne = []
                        for j,v in enumerate(lignes):
                            ligne.append(v.strip() if v else "vide")
                        dataUtile.append(ligne)
                
                else:
                    horairesLigne=[]
                    for i,col in enumerate(ligneWb[3]):
                        if col:
                            horairesLigne.append(col)
                    joursLigne=[]
                    for i,lignes in enumerate(ligneWb[4:]):
                        joursLigne.append(lignes[0].strip())
                    if len(joursLigne) != 6:
                        return Response(
                            {
                                "texte":"Ligne jour manquant !",
                                "aide":"Ligne requis: 'Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'"
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    colUtile=["Horaire"]
                    nbCol=len(ligneWb[3])
                    nbCase=(nbCol-2)//4
                    for j,jour in enumerate(joursLigne):
                        colUtile.append(jour)
                        for k in range(nbCase-1):
                            colUtile.append(f"Unnamed {j*(nbCase-1)+k}")
                    for i,horaire in enumerate(ligneWb[3]):
                        verif = (i<(nbCol//2) and (i-1)%nbCase==0) or (i>(nbCol//2) and (i-2)%nbCase==0)
                        if verif:
                            ligneDonne=[horaire]
                            for lignes in ligneWb[4:]:
                                for k in range(nbCase):
                                    ligneDonne.append(lignes[i+k] if lignes[i+k] else "vide")
                            dataUtile.append(ligneDonne)   

                df=pd.DataFrame(dataUtile, columns=colUtile)

                if not all(col in df.columns for col in colonnes_requis):
                    return Response({"erreur":"Format invalide. Colonne requis: 'Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'"}, status=status.HTTP_400_BAD_REQUEST)
                colonneVerif=[col for col in df.columns if col in colonnes_requis]
                for i,v in enumerate(colonneVerif):
                    if v != colonnes_requis[i]:
                        return Response({"erreur":"Format invalide. Colonne requis: 'Horaire','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'"}, status=status.HTTP_400_BAD_REQUEST)
                        
                dataContenu=[]
                dataTitre=premierLignes.to_dict(orient='records')
                colonnes=df.columns
                
                for i,ligne in df.iterrows():
                    lignes = {"Horaire":ligne['Horaire']}
                    if ligne['Horaire'] != "vide":
                        j=1
                        while j<len(colonnes):
                            jour=colonnes[j]
                            valeurs=[ligne.get(jour)]
                            k=1
                            while j+k< len(colonnes) and colonnes[j+k] not in jours:
                                valeurs.append(ligne.get(colonnes[j+k]))
                                k+=1
                            j+=k
                            lignes[jour]=valeurs
                        dataContenu.append(lignes)
                data={
                    "titre":dataTitre,
                    "contenu":dataContenu,
                    "typeFichier":typeFichier
                }

                serializer=DataSerializer(data=data)
                if serializer.is_valid():
                    donnee = serializer.validated_data
                    edtCrudInstance=ServiceEdtCrud()
                    response = edtCrudInstance.ajoutEdtListeDonnee(donnee,jours)

                    return Response(response['context'], status=response['status'])
                
                return Response(serializer.errors,status=status.HTTP_401_UNAUTHORIZED)
            
            except Exception as e:
                return Response(str(e), status=status.HTTP_400_BAD_REQUEST)
            
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
