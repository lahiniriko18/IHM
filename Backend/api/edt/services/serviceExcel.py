from io import BytesIO

import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


class ServiceCreerExcel:
    def __init__(self, nbCase, titre1, titre2):
        self.nbCase = nbCase
        self.titre1 = titre1
        self.titre2 = titre2

    def modele1(self):
        nbCol = self.nbCase * 6 + 1

        colonnes = [f"Col{i+1}" for i in range(nbCol)]
        lignes = [["" for i in range(nbCol)] for j in range(5)]

        df = pd.DataFrame(lignes, columns=colonnes)
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="modele", startrow=3)

            wb = writer.book
            ws = writer.sheets["modele"]
            titres = ["", f"Emploi du temps du {self.titre2}", f"{self.titre1}"]
            jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]

            bordure = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for i in range(len(df.columns)):
                colStr = get_column_letter(i + 1)
                ws.column_dimensions[colStr].width = 11.5
            for i in range(1, 5):
                ws.row_dimensions[i].height = 24
            for i in range(5, 10):
                if i != 7:
                    ws.row_dimensions[i].height = None

            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ):
                for cell in row:
                    cell.font = Font(name="Times New Roman", size=12)
                    cell.border = bordure
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )

            ligne1 = list(
                ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column)
            )[0]
            ligne3 = list(
                ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=ws.max_column)
            )[0]
            ligne7 = list(
                ws.iter_rows(min_row=7, max_row=7, min_col=1, max_col=ws.max_column)
            )[0]
            for ligne in [ligne1, ligne3, ligne7]:
                for cell in ligne:
                    cell.border = Border()

            for i in range(1, nbCol, self.nbCase):
                col1 = get_column_letter(i + 1)
                col2 = get_column_letter(i + self.nbCase)
                cr = f"{col1}4:{col2}4"
                ws.merge_cells(cr)
                cell = ws[f"{col1}4"]
                cell.value = jours[(i // self.nbCase)]
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for i, val in enumerate(titres):
                derCol = get_column_letter(nbCol)
                ws.merge_cells(f"A{i+1}:{derCol}{i+1}")
                cell = ws[f"A{i+1}"]
                cell.value = val
                cell.alignment = Alignment(horizontal="center", vertical="center")

            h = 8
            for i in range(5, 10):
                if i != 7:
                    cell = ws[f"A{i}"]
                    cell.value = f"{h}h-{h+2}h"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                h += 2

            exempleDonne = [
                ["IG gr1", "Analyse", "Mme Bénedicte", "S 001"],
                ["IG gr2", "Archi", "Mr Haja", "S 008"],
            ]

            cellHoraire = ws["A4"]
            cellHoraire.border = Border()
            cellHoraire.value = ""
            cellHoraire.alignment = Alignment(horizontal="center", vertical="center")

            v = False
            for i in range(5, 10):
                if i != 7:
                    cellExemple = ws[f"B{i}"]
                    cellExemple.value = "\n".join(exempleDonne[int(v)])
                    cellExemple.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )
                    v = not v

        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachement; filename="modele1.xlsx"'
        return response

    def modele2(self):
        nbCol = self.nbCase * 4 + 2

        colonnes = [f"Col{i+1}" for i in range(nbCol)]
        lignes = [["" for i in range(nbCol)] for j in range(6)]

        df = pd.DataFrame(lignes, columns=colonnes)
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="modele", startrow=3)

            wb = writer.book
            ws = writer.sheets["modele"]
            titres = ["", f"Emploi du temps du {self.titre2}", f"{self.titre1}"]
            jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]

            bordure = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for i in range(len(df.columns)):
                colStr = get_column_letter(i + 1)
                if i == (nbCol // 2):
                    ws.column_dimensions[colStr].width = 2
                    continue
                ws.column_dimensions[colStr].width = 11.5
            for i in range(1, 5):
                ws.row_dimensions[i].height = 24
            for i in range(5, 11):
                ws.row_dimensions[i].height = None

            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ):
                for cell in row:
                    cell.font = Font(name="Times New Roman", size=12)
                    cell.border = bordure
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )

            ligne1 = list(
                ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column)
            )[0]
            ligne3 = list(
                ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=ws.max_column)
            )[0]
            for ligne in [ligne1, ligne3]:
                for cell in ligne:
                    cell.border = Border()
            i = 1
            h = 8
            while i < nbCol:
                if i == (nbCol // 2):
                    i += 1
                    h += 2
                    continue
                col1 = get_column_letter(i + 1)
                col2 = get_column_letter(i + self.nbCase)
                cr = f"{col1}4:{col2}4"
                ws.merge_cells(cr)
                cell = ws[f"{col1}4"]
                cell.value = f"{h}h-{h+2}h"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                i += self.nbCase
                h += 2

            for i, val in enumerate(titres):
                derCol = get_column_letter(nbCol)
                ws.merge_cells(f"A{i+1}:{derCol}{i+1}")
                cell = ws[f"A{i+1}"]
                cell.value = val
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for i in range(5, 11):
                cell = ws[f"A{i}"]
                cell.value = jours[i - 5]
                cell.alignment = Alignment(horizontal="center", vertical="center")

            exempleDonne = [
                ["IG gr1", "Analyse", "Mme Bénedicte", "S 001"],
                ["IG gr2", "Archi", "Mr Haja", "S 008"],
            ]

            cellHoraire = ws["A4"]
            cellHoraire.border = Border()
            cellHoraire.value = ""
            cellHoraire.alignment = Alignment(horizontal="center", vertical="center")
            for i in range(5, 11):
                cellExemple = ws[f"B{i}"]
                cellExemple.value = "\n".join(exempleDonne[(i - 1) % 2])
                cellExemple.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

            m = nbCol // 2
            col = get_column_letter(m + 1)
            ws.merge_cells(f"{col}4:{col}10")
            cell = ws[f"{col}4"]
            cell.border = Border()
            cell.value = ""
            cell.alignment = Alignment(horizontal="center", vertical="center")

        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachement; filename="modele2.xlsx"'
        return response

    def modeleEdtProf(self, semaine=None, donnees=None):
        nbCol = 6
        colonnes = [f"Col{i+1}" for i in range(nbCol)]
        lignes = [["" for i in range(nbCol)] for j in range(9)]

        df = pd.DataFrame(lignes, columns=colonnes)
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="modele", startrow=3)

            wb = writer.book
            ws = writer.sheets["modele"]
            titres = ["", f"Emploi du temps du {self.titre2}", f"{self.titre1}"]
            jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]

            bordure = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for i in range(len(df.columns)):
                colStr = get_column_letter(i + 1)
                ws.column_dimensions[colStr].width = 18
            for i in range(1, 5):
                ws.row_dimensions[i].height = 24
            for i in range(5, 10):
                if i != 7:
                    ws.row_dimensions[i].height = None

            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ):
                for cell in row:
                    cell.font = Font(name="Times New Roman", size=12)
                    cell.border = bordure
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )

            ligne1 = list(
                ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column)
            )[0]
            ligne3 = list(
                ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=ws.max_column)
            )[0]
            ligne9 = list(
                ws.iter_rows(min_row=9, max_row=9, min_col=1, max_col=ws.max_column)
            )[0]
            for ligne in [ligne1, ligne3, ligne9]:
                for cell in ligne:
                    cell.border = Border()

            for i in range(nbCol):
                col1 = get_column_letter(i + 1)
                cell = ws[f"{col1}4"]
                cell.value = jours[i]
                cell.alignment = Alignment(horizontal="center", vertical="center")

            derCol = get_column_letter(nbCol)
            ws.merge_cells(f"A9:{derCol}9")
            cell = ws["A9"]
            cell.value = ""
            for i, val in enumerate(titres):
                ws.merge_cells(f"A{i+1}:{derCol}{i+1}")
                cell = ws[f"A{i+1}"]
                cell.value = val
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for date, donnee in donnees.items():
                if date != "professeur":
                    indexDate = semaine.index(date)
                    col = get_column_letter(indexDate + 1)
                    matinSoir = [10, 5]
                    for val in donnee:
                        heureDebut = val["heureDebut"]
                        heureFin = val["heureFin"]
                        indexMs = int(val["matinSoir"])
                        indexLigne = matinSoir[indexMs]
                        d = [val["classe"], val["matiere"], val["salle"]]

                        cellHoraire = ws[f"{col}{indexLigne}"]
                        cellHoraire.value = f"{heureDebut}-{heureFin}"
                        cellHoraire.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

                        cellDonne = ws[f"{col}{indexLigne+1}"]
                        cellDonne.value = "\n".join(d)

                        matinSoir[indexMs] += 2

        buffer.seek(0)
        return buffer
